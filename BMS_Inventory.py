import tkinter as tk
from tkinter import messagebox
import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border, Side
from openpyxl.styles import Alignment
import datetime
import time
import os
from PIL import ImageTk, Image
import sys
from importlib.metadata import version
import png
import tempfile
from subprocess import call
import psycopg2
import hashlib


def add_product(categoria_object, marca_object, modelo_object, serie_object, pais_object, inv_num_object, descripcion_object, precio_base_object, precio_ask_object, tables_list, root_import, quantity_object, user_id):

	categoria = categoria_object.get()
	marca = marca_object.get()
	modelo = modelo_object.get()
	serie = serie_object.get()
	pais = pais_object.get()
	inv_num = inv_num_object.get()
	descripcion = descripcion_object.get('1.0', 'end-1c')
	precio_base = precio_base_object.get()
	precio_ask = precio_ask_object.get()
	quantity = quantity_object.get()

	# Gets the category_id
	query = f"SELECT id FROM individual_category WHERE UPPER(category_name)= UPPER('{categoria}')"
	cur_main.execute(query)
	category_id = cur_main.fetchone()[0]


	if len(precio_base) == 0:
		precio_base = '0'
	if len(precio_ask) == 0:
		precio_ask = '0'

	# Used if multiple quantities of an item are added, but the description and price are the same. Only serial and inv numwould be different for each.
	specs_no_serial_inv = [category_id, marca, modelo, pais, descripcion, precio_base, precio_ask] # Does not include serial nor inv number nor quantity

	# Used if multiple quantities of an item are added and the description, price, and serial and inv are all different for each.
	specs_no_desc_price_serial_inv = [category_id, marca, modelo, pais] # Does not include serial, inv, description, nor price. Also no quantity


	add_prod = True
	add_prod_font = "Calibri 20 bold"
	if len(marca) == 0:
		add_prod = False
		label_warning = tk.Label(root_import, font=add_prod_font, bg='#D9D9D9', fg='red', text="WARNING: Brand is Missing")
	elif len(modelo) == 0:
		add_prod = False
		label_warning = tk.Label(root_import, font=add_prod_font, bg='#D9D9D9', fg='red', text="WARNING: Model is Missing")
	elif len(serie) == 0:
		add_prod = False
		label_warning = tk.Label(root_import, font=add_prod_font, bg='#D9D9D9', fg='red', text="WARNING: Serial # is Missing")
	elif len(pais) == 0:
		add_prod = False
		label_warning = tk.Label(root_import, font=add_prod_font, bg='#D9D9D9', fg='red', text="WARNING: Country is Missing")
	elif len(inv_num) == 0:
		add_prod = False
		label_warning = tk.Label(root_import, font=add_prod_font, bg='#D9D9D9', fg='red', text="WARNING: Inventory # is Missing")
	elif len(descripcion) == 0:
		add_prod = False
		label_warning = tk.Label(root_import, font=add_prod_font, bg='#D9D9D9', fg='red', text="WARNING: Description is Missing")	

	## PRICES DONT MATTER IF THEY ARE EMPTY RIGHT NOW ##
		# elif len(precio_base) == 0:
		# 	add_prod = False
		# 	label_warning = tk.Label(root_import, text="WARNING: Base Price is Missing")
		# elif len(precio_ask) == 0:
		# 	add_prod = False
		# 	label_warning = tk.Label(root_import, text="WARNING: Asking Price is Missing")

	elif (precio_base.replace('.', '', 1).isdigit() == False) and len(precio_base) > 0:
		add_prod = False
		label_warning = tk.Label(root_import, font=add_prod_font, bg='#D9D9D9', fg='red', text="WARNING: Base Price must be a number")
	elif precio_ask.replace('.', '', 1).isdigit() == False and len(precio_ask) > 0:
		add_prod = False
		label_warning = tk.Label(root_import, font=add_prod_font, bg='#D9D9D9', fg='red', text="WARNING: Asking Price must be a number")
	elif len(quantity) == 0:
		add_prod = False
		label_warning = tk.Label(root_import, font=add_prod_font, bg='#D9D9D9', fg='red', text="WARNING: Quantity is Missing")
	elif quantity == '0':
		add_prod = False
		label_warning = tk.Label(root_import, font=add_prod_font, bg='#D9D9D9', fg='red', text="WARNING: Quantity cannot be 0")
	elif (quantity.replace('.', '', 1).isdigit() == False):
		add_prod = False
		label_warning = tk.Label(root_import, font=add_prod_font, bg='#D9D9D9', fg='red', text="WARNING: Quantity must be a number")
	elif (float(quantity).is_integer() == False):
		add_prod = False
		label_warning = tk.Label(root_import, font=add_prod_font, bg='#D9D9D9', fg='red', text="WARNING: Quantity must be a whole number")
	else:
		label_warning = tk.Label(root_import, bg='#D9D9D9')

	label_warning.place(relx=0.3, rely=0.35, relwidth=0.3, relheight=0.06)

	if add_prod == False:
		return

	# Checks if serial number given is already in another item of the same model in the database
	cur_main.execute(f"SELECT * FROM individual_equipment WHERE serial_number = '{serie}' AND UPPER(model) = UPPER('{modelo}')")
	check_serial_existance = cur_main.fetchall()

	# If serial number is already in database, warning is given and function ends, not allowing you to go to the next screen
	if bool(check_serial_existance) == True and (serie != "N/A"):
		add_prod = False
		label_warning = tk.Label(root_import, font=add_prod_font, bg='#D9D9D9', fg='red', text="WARNING: Serial # Already in Inv")
		label_warning.place(relx=0.3, rely=0.35, relwidth=0.3, relheight=0.06)
		if add_prod == False:
			return

	# Checks if there already exists an item with that inventory number
	cur_main.execute(f"SELECT * FROM individual_equipment WHERE inv_num = '{inv_num}'")
	checking_inv_num_existance = cur_main.fetchall()

	if bool(checking_inv_num_existance) == True:
		add_prod = False
		label_warning = tk.Label(root_import, font=add_prod_font, bg='#D9D9D9', fg='red', text="WARNING: Inv # Already in Inv")
		label_warning.place(relx=0.3, rely=0.35, relwidth=0.3, relheight=0.06)
		if add_prod == False:
			return

	print(serie_object.cget("state"))


	# Clears info in entries to allow next item to be added
	categoria_object.set(tables_list[0])
	marca_object.delete(0, 'end')
	modelo_object.delete(0, 'end')
	serie_object.delete(0, 'end')
	pais_object.delete(0, 'end')
	inv_num_object.delete(0, 'end')
	descripcion_object.delete('1.0', 'end')
	precio_base_object.delete(0, 'end')
	precio_ask_object.delete(0, 'end')
	quantity_object.delete(0, 'end')
	quantity_object.insert(0, 1)


	# If quantity being added is great that one
	if int(quantity) > 1:

		serial_state = serie_object.cget('state')

		root_sameordiff = tk.Toplevel()
		root_sameordiff.title('Import2')

		canvas = tk.Canvas(root_sameordiff, width=400, height=100)
		canvas.pack()

		label_question = tk.Label(root_sameordiff, text='Do items have the same price and description?', font='Calibri 12')
		label_question.place(relx=0.1, relwidth=0.8, relheight=0.35)

		# Items have same price and desc. Serial and inv num must be added
		button_yes = tk.Button(root_sameordiff, bg='#B6FE90', text='YES', font='Calibri 12', command=lambda: yes(root_sameordiff, quantity, specs_no_serial_inv, serie, inv_num, user_id, serial_state, modelo, root_import))
		button_yes.place(relx=0.2, rely=0.4, relwidth=0.2, relheight=0.4)

		# Items have different price and desc. Price, description, serial, and inv must be added
		button_no = tk.Button(root_sameordiff, bg='#FE9090', text='NO', font='Calibri 12', command=lambda: no(root_sameordiff, quantity, specs_no_desc_price_serial_inv, serie, inv_num, descripcion, precio_base, precio_ask, user_id, serial_state, modelo, root_import))
		button_no.place(relx=0.6, rely=0.4, relwidth=0.2, relheight=0.4)


	# Quantity = 1
	else:

		insert_query = "INSERT INTO individual_equipment (inv_num, brand, model, country, serial_number, description, base_price, price, import_date, category_id, import_user_id) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
		insert_data = (inv_num, marca, modelo, pais, serie, descripcion, precio_base, precio_ask, datetime.datetime.now(), category_id, user_id)
		print("data:", insert_data)

		cur_main.execute(insert_query, insert_data)
		conn_main.commit()
		tk.messagebox.showinfo("Import", "Item Added Successfully", parent=root_import)


		# entrada_product_excel(category_id, marca, modelo, serie, pais, inv_num, descripcion, precio_base, precio_ask, user_id)


def yes(root_sameordiff, quantity, specs_no_serial_inv, serie, inv_num, user_id, serial_state, model, root_import):
	global current_serial

	root_yes = tk.Toplevel()
	root_yes.title('Serial')

	root_sameordiff.destroy()

	serial_list = []
	inv_num_list = []

	current_serial = 1

	canvas = tk.Canvas(root_yes, width=650, height=225)
	canvas.pack()

	label_serial = tk.Label(root_yes, text='Serial #', font='Calibri 16')
	label_serial.place(relx=0.2, rely=0.15, relwidth=0.6, relheight=0.1)

	label_inv_num = tk.Label(root_yes, text='   Inv #', font='Calibri 16')
	label_inv_num.place(relx=0.2, rely=0.55, relwidth=0.6, relheight=0.15)

	label_outof = tk.Label(root_yes, font="calibri 10", text='%s / %s' %(current_serial, quantity))
	label_outof.place(relx=0.3, rely=0.91, relwidth=0.4, relheight=0.08)

	entry_serial = tk.Entry(root_yes, font='Calibri 16')
	entry_serial.place(relx=0.2, rely=0.25, relwidth=0.6, relheight=0.15)
	entry_serial.insert(0, serie)
	entry_serial.config(state=serial_state)


	entry_inv_num = tk.Entry(root_yes, font="Calibri 16")
	entry_inv_num.place(relx=0.2, rely=0.70, relwidth=0.6, relheight=0.15)
	entry_inv_num.insert(0, inv_num)

	add_prod_font = "Calibri 14 bold"
	label_warning = tk.Label(root_yes, font=add_prod_font, anchor="w", bg='#f0f0f0', fg='red', text="")
	label_warning.place(relwidth=0.45, relheight=0.12)

	button_enter = tk.Button(root_yes, font='Calibri 16', bg='light blue', text='ENTER', command=lambda: next_serial(entry_serial, entry_inv_num, serial_list, inv_num_list, label_outof, quantity, root_yes, specs_no_serial_inv, user_id, serial_state, model, label_warning, root_import))
	button_enter.place(relx=0.825, rely=0.4, relwidth=0.15, relheight=0.35)


def no(root_sameordiff, quantity, specs_no_desc_price_serial_inv, serie, inv_num, descripcion, precio_base, precio_ask, user_id, serial_state, model, root_import):
	global current_serial

	root_no = tk.Toplevel()
	root_no.title('Serial, Desc, and Price')

	root_sameordiff.destroy()

	serial_inv_desc_price_list = []

	current_serial = 1

	canvas = tk.Canvas(root_no, width=800, height=400)
	canvas.pack()

	label_serial = tk.Label(root_no, text='Serial #', font='Calibri 16')
	label_serial.place(relx=0.2, rely=0.02, relwidth=0.6, relheight=0.1)

	label_inv_num = tk.Label(root_no, text='Inv #', font='Calibri 16')
	label_inv_num.place(relx=0.2, rely=0.22, relwidth=0.6, relheight=0.1)

	label_outof = tk.Label(root_no, font="Calibri 12", text='%s / %s' %(current_serial, quantity))
	label_outof.place(relx=0.2, rely=0.91, relwidth=0.6, relheight=0.08)

	entry_serial = tk.Entry(root_no, font='Calibri 16')
	entry_serial.place(relx=0.2, rely=0.12, relwidth=0.6, relheight=0.10)
	entry_serial.insert(0, serie)
	entry_serial.config(state=serial_state)

	entry_inv_num = tk.Entry(root_no, font='Calibri 16')
	entry_inv_num.place(relx=0.2, rely=0.32, relwidth=0.6, relheight=0.10)
	entry_inv_num.insert(0, inv_num)

	label_price_base = tk.Label(root_no, text='Base Price:',font='Calibri 16')
	label_price_base.place(relx=0.2, rely=0.45, relwidth=0.2, relheight=0.10)

	label_price_ask = tk.Label(root_no, text='Asking Price', font='Calibri 16')
	label_price_ask.place(relx=0.6, rely=0.45, relwidth=0.2, relheight=0.10)

	entry_precio_base = tk.Entry(root_no, font='Calibri 16')
	entry_precio_base.place(relx=0.2, rely=0.55, relwidth=0.2, relheight=0.10)
	entry_precio_base.insert(0, precio_base)

	entry_precio_ask = tk.Entry(root_no, font='Calibri 16')
	entry_precio_ask.place(relx=0.6, rely=0.55, relwidth=0.2, relheight=0.10)
	entry_precio_ask.insert(0, precio_ask)

	label_desc = tk.Label(root_no, text='Description:', font='Calibri 16')
	label_desc.place(relx= 0.3, rely=0.71, relwidth=0.4, relheight=0.10)

	entry_desc = tk.Entry(root_no, font='Calibri 16')
	entry_desc.place(relx=0.05, rely=0.81, relwidth=0.9, relheight=0.10)
	entry_desc.insert(0, descripcion)

	add_prod_font = "Calibri 14 bold"
	label_warning = tk.Label(root_no, font=add_prod_font, anchor="w", bg='#f0f0f0', fg='red', text="")
	label_warning.place(relwidth=0.45, relheight=0.12)

	button_enter = tk.Button(root_no, font='Calibri 16', bg='light blue', text='ENTER', command=lambda: next_serial_desc_price(entry_serial, entry_inv_num, entry_desc, entry_precio_base, entry_precio_ask, serial_inv_desc_price_list, label_outof, quantity, root_no, specs_no_desc_price_serial_inv, user_id, serial_state, model, label_warning, root_import))
	button_enter.place(relx=0.825, rely=0.3, relwidth=0.15, relheight=0.4)

	return

def next_serial(serial_num_object, inv_num_object, serial_list, inv_num_list, label_outof, quantity, root_yes, specs_no_serial_inv, user_id, serial_state, model, label_warning, root_import):
	global current_serial

	serial_num = serial_num_object.get()
	inv_num = inv_num_object.get()

	add_prod_font = "Calibri 12 bold"

	# Checks if serial number given is already in another item of the same model in the database
	cur_main.execute(f"SELECT * FROM individual_equipment WHERE serial_number = '{serial_num}' AND UPPER(model) = UPPER('{model}')")
	check_serial_existance = cur_main.fetchall()

	# If serial number is already in database, warning is given and function ends, not allowing you to go to the next screen
	if bool(check_serial_existance) == True and (serial_num != "N/A"):
		add_prod = False
		label_warning.config(text="WARNING: Serial # Already in Inv")
		if add_prod == False:
			return

	# Checks if inventory number given is already i another item of any type in the database
	cur_main.execute(f"SELECT * FROM individual_equipment WHERE inv_num = '{inv_num}'")
	check_inv_num_existance = cur_main.fetchall()

	# If inventory number is already in database, warning is given and function ends, not allowing you to go to the next screen
	if bool(check_inv_num_existance) == True:
		add_prod = False
		label_warning.config(text="WARNING: Inv # Already in Inv")

		if add_prod == False:
			return

	# If serial number is in the list of serial numbers added in current batch, warning label given and function returns
	if (serial_num in serial_list) and (serial_num != "N/A"):
		label_warning.config(text="WARNING: Serial # Already in Inv")
		return

	# If inventory number is in the list of serial numbers added in current batch, warning label given and function returns
	if inv_num in inv_num_list:
		label_warning.config(text="WARNING: Inv # Already in Inv")

		return # If serial number was already typed in in a previous serial slide, it rejects it

	if serial_num == "":
		label_warning.config(text="WARNING: Serial # Cannot be empty")
		return

	if inv_num == "":
		label_warning.config(text="WARNING: Inv # Cannot be empty")
		return

	# If none of the if statements ran, meaning no error, label is removed
	label_warning.config(text="")


	serial_list.append(serial_num)
	inv_num_list.append(inv_num)

	current_serial += 1

	print('serial_list', serial_list)

	if current_serial > int(quantity):
		root_yes.destroy()
		add_product_same_desc_price(quantity, serial_list, inv_num_list, specs_no_serial_inv, user_id, root_import)
	else:	
		inv_num_object.delete(0, 'end')
		if serial_state == "normal":
			serial_num_object.delete(0, 'end')
		label_outof.config(text='%s / %s' %(current_serial, quantity))

def add_product_same_desc_price(quantity, serial_list, inv_num_list, specs_no_serial_inv, user_id, root_import): # This function is for adding products that have the same price and desc
	category_id = specs_no_serial_inv[0]
	marca = specs_no_serial_inv[1]
	modelo = specs_no_serial_inv[2]
	pais = specs_no_serial_inv[3]

	descripcion = specs_no_serial_inv[4]
	precio_base = specs_no_serial_inv[5]
	precio_ask = specs_no_serial_inv[6]


	for i in range(int(quantity)):

		insert_query = "INSERT INTO individual_equipment (inv_num, brand, model, country, serial_number, description, base_price, price, import_date, category_id, import_user_id) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
		insert_data = (inv_num_list[i], marca, modelo, pais, serial_list[i], descripcion, precio_base, precio_ask, datetime.datetime.now(), category_id, user_id)

		cur_main.execute(insert_query, insert_data)
		conn_main.commit()
	tk.messagebox.showinfo("Import", "Items Added Successfully", parent=root_import)

def next_serial_desc_price(serial_num_object, inv_num_object, desc_object, price_base_object, price_ask_object, serial_inv_desc_price_list, label_outof, quantity, root_no, specs_no_desc_price_serial_inv, user_id, serial_state, model, label_warning, root_import):
	global current_serial

	serial_num = serial_num_object.get()
	inv_num = inv_num_object.get()
	description = desc_object.get()
	price_base = price_base_object.get()
	price_ask = price_ask_object.get()

	if len(price_base) == 0:
		price_base = '0'
	if len(price_ask) == 0:
		price_ask = '0'

	add_prod_font = "Calibri 14 bold"

	cur_main.execute(f"SELECT * FROM individual_equipment WHERE serial_number = '{serial_num}' AND model = '{model}'")
	check_serial_existance = cur_main.fetchall()

	cur_main.execute(f"SELECT * FROM individual_equipment WHERE inv_num = '{inv_num}'")
	check_inv_num_existance = cur_main.fetchall()

	# WARNINGS
	if (bool(check_serial_existance) == True) and (serial_num != "N/A"):
		add_prod = False
		label_warning.configure(text="WARNING: Serial # Already in Inv")
		if add_prod == False:
			return

	if bool(check_inv_num_existance) == True:
		add_prod = False
		label_warning.configure(text="WARNING: Inv # Already in Inv")
		if add_prod == False:
			return

	serial_num_found = False
	for l in serial_inv_desc_price_list:
		if serial_num == l[0]:
			serial_num_found = True

	inv_num_found = False
	for l in serial_inv_desc_price_list:
		if inv_num == l[1]:
			inv_num_found = True
		

	if serial_num_found == True:
		label_warning.configure(text="WARNING: Serial # Already in Inv")
		return # If serial number was already typed in in a previous serial slide, it rejects it

	if inv_num_found == True:
		label_warning.configure(text="WARNING: Inv # Already in Inv")
		return # If serial number was already typed in in a previous serial slide, it rejects it

	if inv_num == "":
		label_warning.configure(text="WARNING: Inv # Cannot be empty")
		return

	if serial_num == "":
		label_warning.configure(text="WARNING: Serial # Cannot be empty")
		return

	if (price_base.replace('.', '', 1).isdigit() == False) and (len(price_base) > 0):
		label_warning.configure(text="WARNING: Price must be a number")
		return

	if (price_ask.replace('.', '', 1).isdigit() == False) and (len(price_ask) > 0):
		label_warning.configure(text="WARNING: Price must be a number")
		return		


	# If none of the if statements ran, meaning no error, label is removed
	label_warning.configure(text="")

	serial_inv_desc_price_list.append([serial_num, inv_num, description, price_base, price_ask])

	current_serial += 1

	if current_serial > int(quantity):
		root_no.destroy()
		add_product_diff_desc_price(quantity, serial_inv_desc_price_list, specs_no_desc_price_serial_inv, user_id, root_import)
	else:
		serial_num_object.delete(0, 'end')
		desc_object.delete(0, 'end')
		price_base_object.delete(0, 'end')
		price_ask_object.delete(0, 'end')
		if serial_state == "normal":
			inv_num_object.delete(0, 'end')

		label_outof.config(text='%s / %s' %(current_serial, quantity))

def add_product_diff_desc_price(quantity, serial_inv_desc_price_list, specs_no_desc_price_serial_inv, user_id, root_import): # This function is for adding products with diff desc and prices
	category_id = specs_no_desc_price_serial_inv[0]
	marca = specs_no_desc_price_serial_inv[1]
	modelo = specs_no_desc_price_serial_inv[2]
	pais = specs_no_desc_price_serial_inv[3]


	for i in range(int(quantity)):
		serial_num = serial_inv_desc_price_list[i][0]
		inv_num = serial_inv_desc_price_list[i][1]
		desc = serial_inv_desc_price_list[i][2]
		price_base = serial_inv_desc_price_list[i][3]
		price_ask = serial_inv_desc_price_list[i][4]

		insert_query = "INSERT INTO individual_equipment (inv_num, brand, model, country, serial_number, description, base_price, price, import_date, category_id, import_user_id) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
		insert_data = (serie_num, marca, modelo, pais, serial_num, descripcion, precio_base, precio_ask, datetime.datetime.now(), category_id, user_id)
		conn_main.commit()
	tk.messagebox.showinfo("Import", "Item Added Successfully", parent=root_import)


def import_login():
	root_import_login = tk.Toplevel()
	root_import_login.attributes("-fullscreen", True)

	label_logo = tk.Label(root_import_login, image=photo_logo)
	label_logo.place(relx=0.35, rely=0.1, relwidth=0.3, relheight=0.4)

	login_frame = tk.Frame(root_import_login, borderwidth=5, relief="groove")
	login_frame.place(relx=0.28, rely=0.62, relwidth=0.44, relheight=0.32)

	label_username = tk.Label(root_import_login, font="Calibri 16", anchor='w', text="USERNAME:")
	label_username.place(relx=0.3, rely=0.65, relwidth=0.1, relheight=0.05)

	label_password = tk.Label(root_import_login, font="Calibri 16", anchor='w', text="PASSWORD:")
	label_password.place(relx=0.3, rely=0.71, relwidth=0.1, relheight=0.05)

	entry_username = tk.Entry(root_import_login, font="Calibri 16")
	entry_username.place(relx=0.4, rely=0.65, relwidth=0.3, relheight=0.05)
	entry_username.focus()

	entry_password = tk.Entry(root_import_login, show="*", font="Calibri 16")
	entry_password.place(relx=0.4, rely=0.71, relwidth=0.3, relheight=0.05)

	button_login = tk.Button(root_import_login, font="Calibri 16", text="LOGIN", bg='light gray', command=lambda: import_check_login(entry_username, entry_password, root_import_login))
	button_login.place(relx=0.3, rely=0.81, relwidth=0.3, relheight=0.1)

	button_exit = tk.Button(root_import_login, font="Calibri 16", text="EXIT", bg="pink", command=lambda: exit_window(root_import_login))
	button_exit.place(relx=0.61, rely=0.81, relwidth=0.09, relheight=0.1)

	button_add_user = tk.Button(root_import_login, font="Calibri 16", text="Add user", command=add_user)
	button_add_user.place(relx=0.89, rely=0.01, relwidth=0.1, relheight=0.05)

def add_user():
	window_add_user = tk.Toplevel()
	window_add_user.attributes('-fullscreen', True)


	entry_master = tk.Entry(window_add_user, show="*", font="Calibri 16")
	entry_master.place(relx=0.3, rely=0.05, relwidth=0.4, relheight=0.05)

	label_master = tk.Label(window_add_user, font="Calibri 16", text="Master:")
	label_master.place(relx=0.3, rely=0, relwidth=0.4, relheight=0.05)

	label_username = tk.Label(window_add_user, font="Calibri 16", text="Username:")
	label_username.place(relx=0.2, rely=0.75, relwidth=0.1, relheight=0.05)

	label_password = tk.Label(window_add_user, font="Calibri 16", text="Password:")
	label_password.place(relx=0.2, rely=0.8, relwidth=0.1, relheight=0.05)

	entry_username = tk.Entry(window_add_user, font="Calibri 16")
	entry_username.place(relx=0.3, rely=0.75, relwidth=0.5, relheight=0.05)

	entry_password = tk.Entry(window_add_user, show="*", font="Calibri 16")
	entry_password.place(relx=0.3, rely=0.8, relwidth=0.5, relheight=0.05)

	button_add = tk.Button(window_add_user, font="Calibri 16", bg='light gray', text="Add User", command=lambda: add_user_to_db(entry_master, entry_username, entry_password, window_add_user))
	button_add.place(relx=0.3, rely=0.88, relwidth=0.39, relheight=0.1)

	button_cancel = tk.Button(window_add_user, font="Calibri 16", text="Cancel", bg='pink', command= lambda: exit_window(window_add_user))
	button_cancel.place(relx=0.71, rely=0.88, relwidth=0.09, relheight=0.1)
	return

def add_user_to_db(master_object, username_object, password_object, window):
	master = master_object.get()
	user_username = username_object.get()
	user_password = password_object.get()

	user_password = hashlib.sha256(user_password.encode()).hexdigest()
	cur_main.execute("SELECT password FROM master_password")
	passwords = cur_main.fetchall()
	for password in passwords:
		input_password = password[0]
		if (master == input_password):
			print("User", user_username, "Added.")

			insert_query = "INSERT INTO inv_user (username, password) VALUES (%s, %s)"
			insert_data = (user_username, user_password)

			cur_main.execute(insert_query, insert_data)
			conn_main.commit()
			window.destroy()
			return
		master_object.delete(0, "end")
		tk.messagebox.showinfo("Error", "Master Password Incorrect", parent=window)





def popupmsg(msg):
    popup = tk.Toplevel()
    popup.wm_title("!")
    label = tk.Label(popup, text=msg, font="Helvetica 11")
    label.pack(side="top", fill="x", pady=10)
    B1 = tk.Button(popup, text="Okay", bg='light gray', font="Helvetica 11", command = popup.destroy)
    B1.pack()

def import_check_login(username_object, password_object, root_import_login):

	password = password_object.get()
	username = username_object.get()

	# Gets password from entrybox and hashes it before saving it to a variable
	password = hashlib.sha256(password.encode()).hexdigest()
	print("pass: ", password)

	# Gets if there is any DB with given username and given password (hashed, of course)
	statement = f"SELECT id from inv_user WHERE username='{username}' AND password = '{password}'"
	cur_main.execute(statement)


	user_id_tuple = cur_main.fetchone()
	print(user_id_tuple)

	if user_id_tuple:
		user_id = user_id_tuple[0]
		root_import_login.destroy()
		entrada(user_id)
	else:
		username_object.delete(0, "end")
		password_object.delete(0, "end")
		tk.messagebox.showinfo("Error", "Login Failed", parent=root_import_login)
		# tk.messagebox.showerror(root= root_import_login, title="Wrong Login Info", message="The username or password is incorrect")

	return

def entrada(user_id):
	root_entrada = tk.Toplevel()
	#root_entrada.title('Entrada')
	root_entrada.attributes('-fullscreen', True)

	label_background_entrada = tk.Label(root_entrada, image=photo_entrada)
	label_background_entrada.place(relwidth=1, relheight=1)

	label_quantity = tk.Label(root_entrada, bg='light gray', font="Calibri 16 bold", text='Quantity')
	label_quantity.place(relx=0.08, rely=0.01, relwidth=0.1, relheight=0.05)

	var_quantity = tk.StringVar(root_entrada)
	var_quantity.set(1)

	entry_quantity = tk.Entry(root_entrada, bg='white', font='Calibri 14', textvariable=var_quantity, justify='center')
	entry_quantity.place(relx=0.08, rely=0.07, relwidth=0.1, relheight=0.05)


	### ENTRIES STARTS ###
	query = "SELECT category_name FROM individual_category ORDER BY category_name"
	cur_main.execute(query)
	tables = cur_main.fetchall()

	tables_list = []
	for i in range(len(tables)):
		tables_list.append(tables[i][0].upper())

	entrada_entry_color = 'white'
	entrada_option_color = 'white'

	if len(tables_list) != 0:
		var_type = tk.StringVar(root_entrada)
		var_type.set(tables_list[0])

		optionmenu_tipo = tk.OptionMenu(root_entrada, var_type, *tables_list)
		optionmenu_tipo.place(relx=0.01, rely=0.2425, relwidth=0.28, relheight=0.065)
		optionmenu_tipo.config(bg = entrada_option_color)


	button_create_cat = tk.Button(root_entrada, image= photo_entrada_create_cat, text='Create Catagory', font='Calibri 14', command=lambda:create_cat(root_entrada, user_id))
	button_create_cat.place(relx=0.0225, rely=0.333, relwidth=0.2625, relheight=0.065)

	entry_marca = tk.Entry(root_entrada, bg=entrada_entry_color, font='Calibri 14', justify='center')
	entry_marca.place(relx=0.2975, rely=0.2425, relwidth=0.168, relheight=0.0625)

	entry_modelo = tk.Entry(root_entrada, bg=entrada_entry_color, font='Calibri 14', justify='center')
	entry_modelo.place(relx=0.473, rely= 0.2425, relwidth=0.14, relheight=0.0625)

	entry_serie = tk.Entry(root_entrada, bg=entrada_entry_color, state='normal', font='Calibri 14', justify='center')
	entry_serie.place(relx=0.621, rely=0.2425, relwidth=0.147, relheight=0.0625)

	entry_pais = tk.Entry(root_entrada, bg=entrada_entry_color, font='Calibri 14', justify='center')
	entry_pais.place(relx=0.776, rely=0.2425, relwidth=0.112, relheight=0.0625)

	entry_inv = tk.Entry(root_entrada, bg=entrada_entry_color, font='Calibri 14', justify='center')
	entry_inv.place(relx=0.895, rely=0.2425, relwidth=0.094, relheight=0.0625)

	text_descripcion = tk.Text(root_entrada, font='Calibri 14', wrap='word')
	text_descripcion.place(relx=0.336, rely=0.434, relwidth=0.495, relheight=0.35)

	entry_precio_base = tk.Entry(root_entrada, bg=entrada_entry_color, font='Calibri 14', justify='center')
	entry_precio_base.place(relx=0.375, rely=0.895, relwidth=0.156, relheight=0.05)

	entry_precio_ask = tk.Entry(root_entrada, bg=entrada_entry_color, font='Calibri 14', justify='center')
	entry_precio_ask.place(relx=0.6225, rely=0.895, relwidth=0.156, relheight=0.05)


	### ENTRIES ENDS ###

	button_submit = tk.Button(root_entrada, bg='#8C8187', image=photo_entrada_submit, font='Calibri 16 bold', command=lambda: add_product(var_type, entry_marca, entry_modelo, entry_serie, entry_pais, entry_inv, text_descripcion, entry_precio_base, entry_precio_ask, tables_list, root_entrada, entry_quantity, user_id))
	button_submit.place(relx=0.05, rely=0.825, relwidth=0.225, relheight=0.125)

	button_exit = tk.Button(root_entrada, image=photo_back, bg= 'gray', command=lambda: exit_window(root_entrada))
	button_exit.place(relx=0.86, rely=0.9, relwidth=0.115, relheight=0.06)

	button_serie_toggle = tk.Button(root_entrada, text='Toggle Serial', command=lambda: toggle_serial(entry_serie))
	button_serie_toggle.place(relx=0.627, rely=0.325, relwidth=0.135, relheight=0.0525)

def toggle_serial(entry_object):

	if entry_object.cget('state') == 'normal':
		entry_object.delete(0, 'end')
		entry_object.insert(0, 'N/A')
		entry_object.config(state='disabled')
	else:
		entry_object.config(state='normal')
		entry_object.delete(0, 'end')


def create_cat(root_entrada, user_id):
	root_create_cat = tk.Toplevel()
	root_create_cat.title('Create Catagory')

	canvas = tk.Canvas(root_create_cat, width=400, height=100)
	canvas.pack()

	label_create = tk.Label(root_create_cat, text='Create Catagory:', font='Calibri 24')
	label_create.place(relx=0.2, relwidth=0.6, relheight=0.35)

	entry_create = tk.Entry(root_create_cat, font='Calibri 20')
	entry_create.place(relx=0.2, rely=0.4, relwidth=0.6, relheight=0.4)

	button_confirm = tk.Button(root_create_cat, bg='light green', text='>>', font='Calibri 20', command=lambda: create_cat_2(entry_create.get(), root_entrada, root_create_cat, user_id))
	button_confirm.place(relx=0.85, rely=0.4, relwidth=0.1, relheight=0.4)

def create_cat_2(cat_name, root_entrada, root_create_cat, user_id):

	query = f"SELECT * FROM individual_category WHERE UPPER(category_name) = UPPER('{cat_name}')"
	cur_main.execute(query)
	exists = cur_main.fetchone()

	if exists == None:
		print("Creating New Category...")

		insert_query = "INSERT INTO individual_category(category_name) VALUES (%s)"
		insert_data = (cat_name,)

		print(insert_query)
		print(insert_data)
		cur_main.execute(insert_query, insert_data)
		conn_main.commit()

		root_create_cat.destroy()
		root_entrada.destroy()
		entrada(user_id)

	else:
		print("Category already exists")
		root_create_cat.destroy()
		root_entrada.destroy()
		entrada(user_id)


def export_login():
	root_export_login = tk.Toplevel()
	root_export_login.attributes("-fullscreen", True)

	label_logo = tk.Label(root_export_login, image=photo_logo)
	label_logo.place(relx=0.35, rely=0.1, relwidth=0.3, relheight=0.4)

	login_frame = tk.Frame(root_export_login, borderwidth=5, relief="groove")
	login_frame.place(relx=0.28, rely=0.62, relwidth=0.44, relheight=0.32)

	label_username = tk.Label(root_export_login, font="Calibri 16", anchor='w', text="USERNAME:")
	label_username.place(relx=0.3, rely=0.65, relwidth=0.1, relheight=0.05)

	label_password = tk.Label(root_export_login, font="Calibri 16", anchor='w', text="PASSWORD:")
	label_password.place(relx=0.3, rely=0.71, relwidth=0.1, relheight=0.05)

	entry_username = tk.Entry(root_export_login, font="Calibri 16")
	entry_username.place(relx=0.4, rely=0.65, relwidth=0.3, relheight=0.05)
	entry_username.focus()

	entry_password = tk.Entry(root_export_login, show="*", font="Calibri 16")
	entry_password.place(relx=0.4, rely=0.71, relwidth=0.3, relheight=0.05)

	button_login = tk.Button(root_export_login, font="Calibri 16", text="LOGIN", bg='light gray', command=lambda: export_check_login(entry_username, entry_password, root_export_login))
	button_login.place(relx=0.3, rely=0.81, relwidth=0.3, relheight=0.1)

	button_exit = tk.Button(root_export_login, font="Calibri 16", text="EXIT", bg="pink", command=lambda: exit_window(root_export_login))
	button_exit.place(relx=0.61, rely=0.81, relwidth=0.09, relheight=0.1)

	button_add_user = tk.Button(root_export_login, font="Calibri 16", text="Add user", command=add_user)
	button_add_user.place(relx=0.89, rely=0.01, relwidth=0.1, relheight=0.05)

	

def export_check_login(username_object, password_object, root_export_login):

	# Gets password from entrybox and hashes it before saving it to a variable
	password = password_object.get()
	username = username_object.get()
	password = hashlib.sha256(password.encode()).hexdigest()
	print("pass: ", password)

	# Gets if there is any DB with given username and given password (hashed, of course)
	statement = f"SELECT id from inv_user WHERE username='{username}' AND password = '{password}'"
	cur_main.execute(statement)


	user_id_tuple = cur_main.fetchone()
	print(user_id_tuple)

	if user_id_tuple:
		print("welcome")
		user_id = user_id_tuple[0]
		root_export_login.destroy()
		salida2(user_id)
	else:
		username_object.delete(0, "end")
		password_object.delete(0, "end")
		tk.messagebox.showinfo("Error", "Login Failed", parent=root_export_login)

	return

def on_validate(P):
	global entry_scanner
	global inv_list

	if len(P) == 8:  # The 6th entry is taken up by the 2nd entry widget
		entry_scanner.delete(0, "end")
		entry_scanner.after_idle(lambda: entry_scanner.configure(validate='key'))

		inv_num_listbox.insert('end', P)

		inv_list.append(P)

	return True

def salida2(user_id):
	global entry_scanner
	global inv_num_listbox
	global inv_list

	root_salida2 = tk.Toplevel()
	root_salida2.title('Salida')
	root_salida2.attributes('-fullscreen', True)

	inv_list = []

	entry_scanner = tk.Entry(root_salida2, justify="center", bg="light gray", font="Calibri 60", validate="key")
	entry_scanner['validatecommand'] = (entry_scanner.register(on_validate), '%P')
	entry_scanner.place(relx=0.3, rely=0.04, relwidth=0.4, relheight=0.12)

	entry_scanner.focus()

	inv_num_frame = tk.Frame(root_salida2, bg='light blue')
	inv_num_frame.place(relx=0.3, rely=0.2, relwidth=0.4, relheight=0.76)

	# Creates scrollbar
	inv_num_scrollbar = tk.Scrollbar(inv_num_frame, orient='vertical')

	# Creates listbox
	inv_num_listbox = tk.Listbox(inv_num_frame, justify="center", font='consolas 30' ,yscrollcommand = inv_num_scrollbar.set)

	# Configures scrollbar
	inv_num_scrollbar.config(command=inv_num_listbox.yview)

	# Places listbox and scrollbar on screen
	inv_num_listbox.place(relwidth=1, relheight=1)
	inv_num_scrollbar.pack(side='right', fill='y')

	button_submit = tk.Button(root_salida2, text='Confirm', font='Calibri 20', bg='light green', command=lambda: confirm_export(inv_list, root_salida2, user_id))
	button_submit.place(relx=0.75, rely=0.4, relwidth=0.15, relheight=0.16)

	button_cancel = tk.Button(root_salida2, text='Cancel', font='Calibri 20', bg='pink', command=lambda: exit_confirm_window(root_salida2, user_id))
	button_cancel.place(relx=0.1, rely=0.4, relwidth=0.15, relheight=0.16)

	# mom = "MOM"
	# inv_num_listbox.insert(mom)

	# Allow double-click and Enter to select
	# inv_num_listbox.bind('<Double-Button>', lambda x:selectItem(inv_num_listbox.get('anchor')))
	# inv_num_listbox.bind('<Return>', lambda x:selectItem(inv_num_listbox.get('anchor')))
	# inv_sum_accept = tk.Button(root_salida, image=photo_salida_check, bg='#D9D9D9', borderwidth=0, command=lambda: salida_tipo_send(var_category.get(), root_salida))
	# inv_sum_accept.place(relx=0.475, rely=0.21, relwidth=0.05, relheight=0.06)

def inventory():

	root_inventory = tk.Toplevel()
	root_inventory.title('Inventory')
	root_inventory.attributes('-fullscreen', True)

	inventory_frame = tk.Frame(root_inventory, bg='pink')
	inventory_frame.place(relx=0.1, rely=0.3, relwidth=0.8, relheight=0.65)

	# Creates scrollbar
	info_scrollbar = tk.Scrollbar(inventory_frame, orient='vertical')

	# Creates listbox
	info_listbox = tk.Listbox(inventory_frame, font='consolas 18', yscrollcommand = info_scrollbar.set)

	# Configures scrollbar
	info_scrollbar.config(command=info_listbox.yview)

	# Places listbox and scrollbar on screen
	info_listbox.place(relwidth=1, relheight=1)
	info_scrollbar.pack(side='right', fill='y')

	# Allow double-click and Enter to select
	info_listbox.bind('<Double-Button>', lambda x:select_item(info_listbox.get('anchor')))
	info_listbox.bind('<Return>', lambda x:select_item(info_listbox.get('anchor')))

	title_label = tk.Label(root_inventory, anchor='w', font= "consolas 18", text='{:<12}|{:<22}|{:<20}|{:<16}|{:<16}|'.format("Inv Number", "Brand", "Model", "Serial", "Category"))
	title_label.place(relx=0.1, rely=0.25, relwidth=0.8, relheight=0.05)

	populate_inventory("", info_listbox)

	label_inv_num = tk.Label(root_inventory, font='consolas 18', text="Inv #:")
	label_inv_num.place(relx=0.15, rely=0.02, relwidth=0.1, relheight=0.04)

	label_brand = tk.Label(root_inventory, font='consolas 18', text="Brand:")
	label_brand.place(relx=0.45, rely=0.02, relwidth=0.1, relheight=0.04)

	label_model = tk.Label(root_inventory, font='consolas 18', text="Model:")
	label_model.place(relx=0.75, rely=0.02, relwidth=0.1, relheight=0.04)

	label_serial = tk.Label(root_inventory, font='consolas 18', text="Serial:")
	label_serial.place(relx=0.15, rely=0.12, relwidth=0.1, relheight=0.04)

	label_category = tk.Label(root_inventory, font='consolas 18', text="Category:")
	label_category.place(relx=0.75, rely=0.12, relwidth=0.1, relheight=0.04)


	entry_inv_num = tk.Entry(root_inventory, font="consolas 18")
	entry_inv_num.place(relx=0.1, rely=0.06, relwidth=0.2, relheight=0.04)
	entry_inv_num.focus()

	entry_brand = tk.Entry(root_inventory, font="consolas 18")
	entry_brand.place(relx=0.4, rely=0.06, relwidth=0.2, relheight=0.04)

	entry_model = tk.Entry(root_inventory, font="consolas 18")
	entry_model.place(relx=0.7, rely=0.06, relwidth=0.2, relheight=0.04)

	entry_serial = tk.Entry(root_inventory, font="consolas 18")
	entry_serial.place(relx=0.1, rely=0.16, relwidth=0.2, relheight=0.04)

	query = "SELECT category_name FROM individual_category ORDER BY category_name"
	cur_main.execute(query)
	tables = cur_main.fetchall()

	tables_list = [None]
	for i in range(len(tables)):
		tables_list.append(tables[i][0].upper())

	entrada_entry_color = 'white'
	entrada_option_color = 'white'

	if len(tables_list) != 0:
		var_category = tk.StringVar(root_inventory)
		var_category.set(tables_list[0])

		optionmenu_category = tk.OptionMenu(root_inventory, var_category, *tables_list)
		optionmenu_category.place(relx=0.7, rely=0.16, relwidth=0.2, relheight=0.05)
		optionmenu_category.config(bg = entrada_option_color)



	button_search = tk.Button(root_inventory, font="consolas 18", bg="gray", image=photo_search2, command=lambda: search_inventory(entry_inv_num, entry_brand, entry_model, entry_serial, var_category, info_listbox, tables_list))
	button_search.place(relx=0.45, rely=0.14, relwidth=0.1, relheight=0.045)


	button_exit = tk.Button(root_inventory, bg='white', anchor="center", image= photo_back6,command=lambda: exit_window(root_inventory))
	button_exit.place(relx=0.91, rely=0.89, relwidth=0.08, relheight=0.06)

def select_item(selection):
	print("SEL: ", selection)
	print("Feature not yet available")
	inv_num = selection.split()[0]


	query = f"SELECT inv_num, brand, model, country, serial_number, description, base_price, price, import, category_id, import_user_id FROM individual_equipment WHERE inv_num = '{inv_num}'"

	cur_main.execute(query)

	

def search_inventory(inv_num_object, brand_object, model_object, serial_number_object, category_object, listbox_inv, tables_list):

	category = category_object.get()
	query = f"SELECT id FROM individual_category WHERE UPPER(category_name)= UPPER('{category}')"
	cur_main.execute(query)

	if category == "None":
		category_id = None
	else:
		category_id = cur_main.fetchone()[0]


	field_data = []
	field_names = ["inv_num", "brand", "model", "serial_number", "category_id"]

	# Adds user-inputted info to a list. This will be used to create query used to search for records with given parameters

	field_data.append(inv_num_object.get())
	field_data.append(brand_object.get())
	field_data.append(model_object.get())
	field_data.append(serial_number_object.get())
	field_data.append(str(category_id))


	# Clears entry fields after searching
	inv_num_object.delete(0,'end')
	brand_object.delete(0,'end')
	model_object.delete(0,'end')
	serial_number_object.delete(0,'end')
	category_object.set(tables_list[0])

	query_conditions_list = []
	for index in range(len(field_data)):
		if field_data[index] and field_data[index] != "None":
			print("FIELD DATA:", field_data[index])
			# If search parameter is a digit (hospital, company ID or any other digit data), it does not user "UPPER" in query"
			if field_data[index].isdigit():
				query_conditions_list.append( f"{field_names[index]} = '{field_data[index]}'")
			else:
				query_conditions_list.append( f"UPPER({field_names[index]}) = UPPER('{field_data[index]}')")
	
	query_conditions_string = " AND ".join(query_conditions_list)
	
	if query_conditions_string:
		update_listbox(query_conditions_string, listbox_inv)
	else:
		update_listbox("", listbox_inv)


def update_listbox(conditions, listbox_inv):
	delete_listbox(listbox_inv)
	populate_inventory(conditions, listbox_inv)
	return

def delete_listbox(listbox_inv):
	listbox_inv.delete(0,'end')
	return

def populate_inventory(conditions, info_listbox):
	if conditions:
		query = "SELECT inv_num, brand, model, serial_number, category_id FROM individual_equipment WHERE export_date is null AND " + conditions
	else:
		query = "SELECT inv_num, brand, model, serial_number, category_id, export_date FROM individual_equipment WHERE export_date is null"


	cur_main.execute(query)
	equipment_list = cur_main.fetchall()



	for equipment in equipment_list:
		inv_num = shorten_display(equipment[0], 12)
		brand = shorten_display(equipment[1], 22)
		model = shorten_display(equipment[2], 20)
		serial_num = shorten_display(equipment[3], 16)

		category_id = equipment[4]
		statement = f"SELECT category_name FROM individual_category WHERE id='{category_id}'"
		cur_main.execute(statement)		
		category = cur_main.fetchone()[0]

		category = shorten_display(category, 16)

		info_listbox.insert('end', '{:<12} {:<22} {:<20} {:<16} {:<16}'.format(inv_num, brand, model, serial_num, category))

def shorten_display(string, length):
	'''Given a string and a length, it shortens the word to length,
	   with last three characters being dots (...)'''
	
	string = str(string)

	if len(string) <= length:
		return string.upper()

	string = string[:length-3]
	string += '...'

	return string.upper()


def history():

	root_history = tk.Toplevel()
	#root_history.attributes('-fullscreen', True)

	canvas = tk.Canvas(root_history, bg='light blue', width=500, height=150).pack()

	label_month = tk.Label(root_history, bg="light blue", font="Calibri 16", text="MM")
	label_month.place(relx=0.3, rely=0.01, relwidth=0.08, relheight=0.3)

	label_day = tk.Label(root_history, bg="light blue", font="Calibri 16", text="DD")
	label_day.place(relx=0.41, rely=0.01, relwidth=0.08, relheight=0.3)

	label_year = tk.Label(root_history, bg="light blue", font="Calibri 16", text="YYYY")
	label_year.place(relx=0.52, rely=0.01, relwidth=0.16, relheight=0.3)

	entry_month = tk.Entry(root_history, font="Calibri 20")
	entry_month.place(relx=0.3, rely=0.3, relwidth=0.08, relheight=0.3)
	entry_month.focus()

	label_slash1 = tk.Label(root_history, bg="light blue", font="Calibri 22", text="/")
	label_slash1.place(relx=0.38, rely=0.3, relwidth=0.03, relheight=0.3)

	entry_day = tk.Entry(root_history, font="Calibri 20")
	entry_day.place(relx=0.41, rely=0.3, relw=0.08, relheight=0.3)

	label_slash2 = tk.Label(root_history, bg="light blue", font="Calibri 22", text="/")
	label_slash2.place(relx=0.49, rely=0.3, relwidth=0.03, relheight=0.3)

	entry_year = tk.Entry(root_history, font="Calibri 20")
	entry_year.place(relx=0.52, rely=0.3, relw=0.16, relheight=0.3)

	button_hist_search = tk.Button(root_history, text='Search', command=lambda: history_search(int(entry_month.get()), int(entry_day.get()), int(entry_year.get()), root_history))
	button_hist_search.place(relx=0.4, rely=0.65, relwidth=0.2, relheight=0.3)


	#path = os.path.join('entradas_y_salidas', 'Entradas_Salidas.xlsx')
	#os.startfile(path)

	return

def history_search(month, day, year, root):
	#create a copy of excel sheet of that month/year
	#open up the copy for user to see
	#when program finishes, it will attempt to delete folder in which the copy was saved

	root_history_search = tk.Toplevel()
	root_history_search.attributes("-fullscreen", True)

	root.destroy()

	try:
		day_of_interest = datetime.datetime(year, month, day, 23, 59, 59)
	except:
		print("wat")
	title_label = tk.Label(root_history_search, anchor='w', font= "consolas 18", text='{:<12}|{:<22}|{:<20}|{:<16}|{:<16}|'.format("Inv Number", "Brand", "Model", "Serial", "Category"))
	title_label.place(relx=0.1, rely=0.15, relwidth=0.8, relheight=0.05)

	inventory_frame = tk.Frame(root_history_search, bg='pink')
	inventory_frame.place(relx=0.1, rely=0.2, relwidth=0.8, relheight=0.65)

	label_history_inv = tk.Label(root_history_search, font="Calibri 24 bold", text="Inventory On:")
	label_history_inv.place(relx=0.3, rely=0, relwidth=0.4, relheight=0.06)


	label_date = tk.Label(root_history_search, font="Calibri 18", text=f'{month}/{day}/{year}')
	label_date.place(relx=0.3, rely=0.06, relwidth=0.4, relheight=0.06)
	# Creates scrollbar
	history_scrollbar = tk.Scrollbar(inventory_frame, orient='vertical')

	# Creates listbox
	history_listbox = tk.Listbox(inventory_frame, font='consolas 18', yscrollcommand = history_scrollbar.set)

	# Configures scrollbar
	history_scrollbar.config(command=history_listbox.yview)

	# Places listbox and scrollbar on screen
	history_listbox.place(relwidth=1, relheight=1)
	history_scrollbar.pack(side='right', fill='y')

	# Allow double-click and Enter to select
	history_listbox.bind('<Double-Button>', lambda x:selectItem(history_listbox.get('anchor')))
	history_listbox.bind('<Return>', lambda x:selectItem(history_listbox.get('anchor')))

	conditions = f"(import_date <= '{day_of_interest}') AND ( (export_date > '{day_of_interest}') OR (export_date is NULL) )"

	populate_inventory(conditions, history_listbox)

	button_back = tk.Button(root_history_search, bg="light gray", font="Calibri 16", text="Back", command= lambda: exit_window(root_history_search))
	button_back.place(relx=0.3, rely=0.875, relwidth=0.1, relheight=0.1)

	button_download = tk.Button(root_history_search, bg="light gray", font="Calibri 16", text="Download", command = lambda: history_download(conditions, year, month, day))
	button_download.place(relx=0.6, rely=0.875, relwidth=0.1, relheight=0.1)

def history_download(conditions, year, month, day):


	file_name = str(year) + "_" + str(month) + '_' + str(day) + '.xlsx'

	full_dir_name = os.path.join('history', file_name)

	# Attempts to open excel sheet. If it does not exist, it creates one
	inventory_database = openpyxl.Workbook()
	sheet = inventory_database.active


	sheet.cell(row=1, column=1, value = 'Inv #')
	sheet.cell(row=1, column=2, value = 'Brand')
	sheet.cell(row=1, column=3, value = 'Model')
	sheet.cell(row=1, column=4, value = 'Serial #')
	sheet.cell(row=1, column=5, value = 'Category')
	sheet.cell(row=1, column=6, value = 'Country')

	sheet.cell(row=1, column=7, value = 'Base Price')
	sheet.cell(row=1, column=8, value = 'Asking Price')

	sheet.cell(row=1, column=9, value = 'Description')

	# Create query that selects all items that meet condition. 
	query = "SELECT inv_num, brand, model, serial_number, category_id, country, base_price, price, description FROM individual_equipment WHERE " + conditions
	cur_main.execute(query)
	equipment_list = cur_main.fetchall()





	# Info begins to be added onto the excel sheet
	ROW = 2

	empty_row = sheet.cell(row=ROW, column=1).value

	query = "SELECT inv_num, brand, model, serial_number, category_id, country, base_price, price, description FROM individual_equipment WHERE " + conditions
	for equipment in equipment_list:

		inv_number = equipment[0]
		brand = equipment[1]
		model = equipment[2]
		serial_num = equipment[3]
		category_id = equipment[4]
		country = equipment[5]
		base_price = equipment[6]
		price = equipment[7]
		description = equipment[8]

		query = f"SELECT category_name FROM individual_category WHERE id = '{category_id}'"
		cur_main.execute(query)
		category_name = cur_main.fetchone()[0]

		sheet.cell(row=ROW, column=1, value = inv_number)
		sheet.cell(row=ROW, column=2, value = brand)
		sheet.cell(row=ROW, column=3, value = model)
		sheet.cell(row=ROW, column=4, value = serial_num)
		sheet.cell(row=ROW, column=5, value = category_name)
		sheet.cell(row=ROW, column=6, value = country)
		sheet.cell(row=ROW, column=7, value = base_price)
		sheet.cell(row=ROW, column=8, value = price)
		sheet.cell(row=ROW, column=9, value = description)

		ROW += 1

	if os.path.exists(full_dir_name):
	  os.remove(full_dir_name)

	inventory_database.save(full_dir_name)

	os.startfile(full_dir_name)


def confirm_export(list_inv_nums, window, user_id):
	export_time = datetime.datetime.now()

	for i in range(len(list_inv_nums)):
		statement = f"UPDATE individual_equipment SET export_date = '{export_time}' WHERE inv_num = '{list_inv_nums[i]}'"
		cur_main.execute(statement)

		statement = f"UPDATE individual_equipment SET export_user_id = '{user_id}' WHERE inv_num = '{list_inv_nums[i]}'"

	conn_main.commit()
	tk.messagebox.showinfo("Export", "Export Successfull!")
	window.destroy()


def exit_window(root):
	root.destroy()

def exit_window_main(root):
	cur_main.close()
	conn_main.close()
	root.destroy()


def exit_confirm_window(root, user_id):
	root.destroy()

def connect_to_database():
	global cur_main
	global conn_main
	conn_main = None

	info_testing = ['hF8$!nfshGAgxPch', 'localhost', 'disable']
	info_production = ['qF7gk3JFrE7Do49z6T', '23.239.24.84', 'require']

	# Change this to move fromt esting to production and vice versa
	info = info_production


	# In PostgreSQL, default username is 'postgres' and password is 'postgres'.
	# And also there is a default database exist named as 'postgres'.
	# Default host is 'localhost' or '127.0.0.1'
	# And default port is '54322'.
	main_database_name = "bms_inventory"
	with open("postgres_user.txt") as f:
		postgres_user = f.readline().rstrip()

	if (info[1] == 'localhost'):
		postgres_user = 'postgres'
	# Connects to main database
	conn_main = psycopg2.connect(
		dbname=main_database_name,
		user=postgres_user,
		password=info[0],
		host=info[1],
		port='5432',
		sslmode=info[2])
	conn_main.autocommit = False


	# Creates cursor for main database
	cur_main = conn_main.cursor()


root = tk.Tk()

root.title('Best Medical Supply Inventory')
root.attributes('-fullscreen', True)
root.iconbitmap("logo.ico")

canvas = tk.Canvas(root, width=1000, height=600)
canvas.pack()

connect_to_database()

#####################################################################
##########################  PHOTOS  #################################
#####################################################################

photo_main = ImageTk.PhotoImage(file = 'backgrounds//main3.jpeg')
photo_main_import = ImageTk.PhotoImage(file = 'buttons//main_import3.jpeg')
photo_main_export = ImageTk.PhotoImage(file = 'buttons//main_export3.jpeg')
photo_main_inventory = ImageTk.PhotoImage(file = 'buttons//main_inventory5.jpeg')
photo_main_history = ImageTk.PhotoImage(file = 'buttons//main_history2.jpeg')
photo_main_shutdown = ImageTk.PhotoImage(file = 'buttons//shutdown2.png')

photo_entrada = ImageTk.PhotoImage(file = 'backgrounds//entrada1.jpeg')
photo_entrada_create_cat = ImageTk.PhotoImage(file= 'buttons//entrada_create_category6.jpeg')
photo_entrada_submit = ImageTk.PhotoImage(file = 'buttons//entrada_submit1.jpeg')
photo_back = ImageTk.PhotoImage(file = 'buttons//back2.jpg')

photo_inv = ImageTk.PhotoImage(file = 'backgrounds//inventory1.jpeg')
photo_inv_searchtype = ImageTk.PhotoImage(file = 'buttons//inv_searchtype3.jpeg')
photo_inv_searchserial = ImageTk.PhotoImage(file = 'buttons//inv_searchserial3.jpeg')
photo_inv_search = ImageTk.PhotoImage(file = 'buttons//search.jpeg')
photo_back2 = ImageTk.PhotoImage(file = 'buttons//back5.jpg')
photo_back6 = ImageTk.PhotoImage(file = 'buttons//back6.jpg')
photo_search2 = ImageTk.PhotoImage(file = 'buttons//search2.png')

photo_salida = ImageTk.PhotoImage(file = 'backgrounds//export.jpeg')
photo_salida_equiptype = ImageTk.PhotoImage(file = 'labels//equip_type2.jpg')
photo_salida_model = ImageTk.PhotoImage(file = 'labels//model3.jpg')
photo_salida_serial = ImageTk.PhotoImage(file = 'labels//serial7.jpg')
photo_salida_submit = ImageTk.PhotoImage(file = 'labels//submit5.jpg')
photo_salida_check = ImageTk.PhotoImage(file = 'buttons//check.png')



# Images 
# image1 = Image.open("logo2.png")
# resized_image = image1.resize((150, 150))
photo_logo = ImageTk.PhotoImage(file = 'logo2.png')


#####################################################################
#####################################################################
#####################################################################


label_background = tk.Label(root, bg='red', image=photo_main)
label_background.place(relwidth=1, relheight=1)

button_entrada = tk.Button(root, text='IMPORT', image=photo_main_import, bg= 'white', font='Calibri 24', command=import_login)
button_entrada.place(relx=0.1, rely=0.25, relwidth=0.3, relheight=0.35)

button_salida = tk.Button(root, text='EXPORT', image=photo_main_export, anchor='n', bg= 'white', font='Calibri 24', command=export_login)
button_salida.place(relx=0.6, rely=0.25, relwidth=0.3, relheight=0.35)

button_inventory = tk.Button(root, text='INVENTORY', image=photo_main_inventory, bg= 'white', font='Calibri 24', command=inventory)
button_inventory.place(relx=0.3, rely=0.7, relwidth=0.4, relheight=0.1)

button_history = tk.Button(root, text='HISTORY', image=photo_main_history, bg= 'white', font='Calibri 24', command=history)
button_history.place(relx=0.3, rely=0.85, relwidth=0.4, relheight=0.1)

button_main_shutdown = tk.Button(root, bg='white', image=photo_main_shutdown, borderwidth=0, command=lambda: exit_window_main(root))
button_main_shutdown.place(relx=0.47, rely=0.1, relwidth=0.06, relheight=0.1)


root.mainloop()


# Entrada shouldnt have a description. That can be done when it is sold.





# BUGS # 

# Make the "Back" Button on the Salidas page clear the salidas inventory list, the price list, the num of items, and the catagory list.
# fix the layout of customer info. Make it so one can review products they are taking out of inventory. Make it look nice.

# QR Code doesnt update the prices or descriptions. Fix it so it does without having to refresh the page
